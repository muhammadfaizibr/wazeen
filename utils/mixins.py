from django.db import models
from django.utils import timezone
from django.contrib.auth import get_user_model
from rest_framework import status
from rest_framework.response import Response
from rest_framework.decorators import action
from django.core.exceptions import PermissionDenied
from django.db.models import Q
from django.utils.translation import get_language


User = get_user_model()


class TimestampMixin(models.Model):
    """Add created_at and updated_at timestamps to models"""
    created_at = models.DateTimeField(auto_now_add=True)
    updated_at = models.DateTimeField(auto_now=True)
    
    class Meta:
        abstract = True


class UserTrackingMixin(models.Model):
    """Track created_by and updated_by users"""
    created_by = models.ForeignKey(
        User, 
        on_delete=models.SET_NULL, 
        null=True, 
        blank=True,
        related_name='%(class)s_created'
    )
    updated_by = models.ForeignKey(
        User, 
        on_delete=models.SET_NULL, 
        null=True, 
        blank=True,
        related_name='%(class)s_updated'
    )
    
    class Meta:
        abstract = True


class SoftDeleteMixin(models.Model):
    """Add soft delete functionality"""
    is_deleted = models.BooleanField(default=False)
    deleted_at = models.DateTimeField(null=True, blank=True)
    deleted_by = models.ForeignKey(
        User, 
        on_delete=models.SET_NULL, 
        null=True, 
        blank=True,
        related_name='%(class)s_deleted'
    )
    
    class Meta:
        abstract = True
    
    def delete(self, using=None, keep_parents=False, user=None):
        """Override delete to perform soft delete"""
        self.is_deleted = True
        self.deleted_at = timezone.now()
        if user:
            self.deleted_by = user
        self.save(using=using)
    
    def hard_delete(self, using=None, keep_parents=False):
        """Perform actual delete"""
        super().delete(using=using, keep_parents=keep_parents)
    
    def restore(self):
        """Restore soft deleted object"""
        self.is_deleted = False
        self.deleted_at = None
        self.deleted_by = None
        self.save()


class SoftDeleteManager(models.Manager):
    """Manager to exclude soft deleted objects by default"""
    
    def get_queryset(self):
        return super().get_queryset().filter(is_deleted=False)
    
    def all_with_deleted(self):
        return super().get_queryset()
    
    def deleted_only(self):
        return super().get_queryset().filter(is_deleted=True)


class ArchiveMixin(models.Model):
    """Add archive functionality"""
    is_archived = models.BooleanField(default=False)
    archived_at = models.DateTimeField(null=True, blank=True)
    archived_by = models.ForeignKey(
        User, 
        on_delete=models.SET_NULL, 
        null=True, 
        blank=True,
        related_name='%(class)s_archived'
    )
    
    class Meta:
        abstract = True
    
    def archive(self, user=None):
        """Archive the object"""
        self.is_archived = True
        self.archived_at = timezone.now()
        if user:
            self.archived_by = user
        self.save()
    
    def unarchive(self):
        """Unarchive the object"""
        self.is_archived = False
        self.archived_at = None
        self.archived_by = None
        self.save()


class ActiveManager(models.Manager):
    """Manager to get only active (non-archived, non-deleted) objects"""
    
    def get_queryset(self):
        qs = super().get_queryset()
        if hasattr(self.model, 'is_archived'):
            qs = qs.filter(is_archived=False)
        if hasattr(self.model, 'is_deleted'):
            qs = qs.filter(is_deleted=False)
        return qs


class MultilingualMixin(models.Model):
    """Add multilingual support for text fields"""
    
    class Meta:
        abstract = True
    
    def get_localized_field(self, field_name, language=None):
        """Get localized version of a field"""
        if not language:
            language = get_language()
        
        # Try to get language-specific field (e.g., name_ar for Arabic)
        lang_code = language.split('-')[0]  # Get language code without region
        localized_field = f"{field_name}_{lang_code}"
        
        if hasattr(self, localized_field):
            localized_value = getattr(self, localized_field)
            if localized_value:
                return localized_value
        
        # Fallback to default field
        return getattr(self, field_name, '')


class OrderableMixin(models.Model):
    """Add ordering functionality"""
    order = models.PositiveIntegerField(default=0)
    
    class Meta:
        abstract = True
        ordering = ['order', 'id']
    
    def move_up(self):
        """Move item up in order"""
        try:
            prev_item = self.__class__.objects.filter(
                order__lt=self.order
            ).order_by('-order').first()
            
            if prev_item:
                prev_item.order, self.order = self.order, prev_item.order
                prev_item.save()
                self.save()
        except Exception:
            pass
    
    def move_down(self):
        """Move item down in order"""
        try:
            next_item = self.__class__.objects.filter(
                order__gt=self.order
            ).order_by('order').first()
            
            if next_item:
                next_item.order, self.order = self.order, next_item.order
                next_item.save()
                self.save()
        except Exception:
            pass


class AuditLogMixin:
    """Mixin for views to log user actions"""
    
    def log_action(self, action, obj=None, extra_data=None):
        """Log user action for audit trail"""
        from django.contrib.contenttypes.models import ContentType
        
        # This would integrate with your audit logging system
        # For now, just a placeholder implementation
        log_data = {
            'user': self.request.user,
            'action': action,
            'timestamp': timezone.now(),
            'ip_address': self.get_client_ip(),
            'user_agent': self.request.META.get('HTTP_USER_AGENT', ''),
        }
        
        if obj:
            log_data.update({
                'content_type': ContentType.objects.get_for_model(obj),
                'object_id': obj.pk,
                'object_repr': str(obj),
            })
        
        if extra_data:
            log_data.update(extra_data)
        
        # TODO: Implement actual audit logging
        # AuditLog.objects.create(**log_data)
    
    def get_client_ip(self):
        """Get client IP address"""
        x_forwarded_for = self.request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]
        else:
            ip = self.request.META.get('REMOTE_ADDR')
        return ip


class RoleBasedViewMixin:
    """Mixin to handle role-based access control"""
    
    allowed_roles = []  # Override in subclasses
    
    def dispatch(self, request, *args, **kwargs):
        """Check role-based access before processing request"""
        if not request.user.is_authenticated:
            return self.handle_no_permission()
        
        if self.allowed_roles and request.user.role not in self.allowed_roles:
            raise PermissionDenied("You don't have permission to access this resource.")
        
        return super().dispatch(request, *args, **kwargs)


class BulkOperationsMixin:
    """Add bulk operations to viewsets"""
    
    @action(detail=False, methods=['post'])
    def bulk_delete(self, request):
        """Bulk delete objects"""
        ids = request.data.get('ids', [])
        if not ids:
            return Response(
                {'error': 'No IDs provided'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        queryset = self.get_queryset().filter(id__in=ids)
        count = queryset.count()
        
        # Perform soft delete if supported
        if hasattr(self.get_serializer().Meta.model, 'is_deleted'):
            queryset.update(
                is_deleted=True,
                deleted_at=timezone.now(),
                deleted_by=request.user
            )
        else:
            queryset.delete()
        
        return Response({
            'message': f'Successfully deleted {count} items'
        })
    
    @action(detail=False, methods=['post'])
    def bulk_update(self, request):
        """Bulk update objects"""
        ids = request.data.get('ids', [])
        updates = request.data.get('updates', {})
        
        if not ids or not updates:
            return Response(
                {'error': 'IDs and updates are required'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Remove protected fields
        protected_fields = ['id', 'created_at', 'created_by']
        for field in protected_fields:
            updates.pop(field, None)
        
        # Add updated_by if supported
        if hasattr(self.get_serializer().Meta.model, 'updated_by'):
            updates['updated_by'] = request.user
        
        queryset = self.get_queryset().filter(id__in=ids)
        count = queryset.update(**updates)
        
        return Response({
            'message': f'Successfully updated {count} items'
        })


class ExportMixin:
    """Add export functionality to viewsets"""
    
    @action(detail=False, methods=['get'])
    def export(self, request):
        """Export data in various formats"""
        format_type = request.query_params.get('format', 'csv')
        
        queryset = self.filter_queryset(self.get_queryset())
        
        if format_type == 'csv':
            return self.export_csv(queryset)
        elif format_type == 'excel':
            return self.export_excel(queryset)
        elif format_type == 'json':
            return self.export_json(queryset)
        else:
            return Response(
                {'error': 'Unsupported format'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
    
    def export_csv(self, queryset):
        """Export as CSV"""
        import csv
        from django.http import HttpResponse
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{self.get_export_filename()}.csv"'
        
        writer = csv.writer(response)
        
        # Write header
        fields = self.get_export_fields()
        writer.writerow(fields)
        
        # Write data
        for obj in queryset:
            row = []
            for field in fields:
                value = self.get_field_value(obj, field)
                row.append(str(value) if value is not None else '')
            writer.writerow(row)
        
        return response
    
    def export_excel(self, queryset):
        """Export as Excel"""
        try:
            import openpyxl
            from django.http import HttpResponse
            
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            
            # Write header
            fields = self.get_export_fields()
            for col, field in enumerate(fields, 1):
                worksheet.cell(row=1, column=col, value=field)
            
            # Write data
            for row, obj in enumerate(queryset, 2):
                for col, field in enumerate(fields, 1):
                    value = self.get_field_value(obj, field)
                    worksheet.cell(row=row, column=col, value=value)
            
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{self.get_export_filename()}.xlsx"'
            
            workbook.save(response)
            return response
            
        except ImportError:
            return Response(
                {'error': 'Excel export requires openpyxl package'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def export_json(self, queryset):
        """Export as JSON"""
        from django.http import JsonResponse
        
        serializer = self.get_serializer(queryset, many=True)
        
        response = JsonResponse(serializer.data, safe=False)
        response['Content-Disposition'] = f'attachment; filename="{self.get_export_filename()}.json"'
        
        return response
    
    def get_export_filename(self):
        """Get filename for export"""
        return f"{self.get_serializer().Meta.model._meta.model_name}s_{timezone.now().strftime('%Y%m%d_%H%M%S')}"
    
    def get_export_fields(self):
        """Get fields to export"""
        return [field.name for field in self.get_serializer().Meta.model._meta.fields]
    
    def get_field_value(self, obj, field_name):
        """Get field value for export"""
        try:
            value = getattr(obj, field_name)
            if hasattr(value, 'all'):  # Many-to-many field
                return ', '.join(str(item) for item in value.all())
            return value
        except AttributeError:
            return None


class SearchMixin:
    """Enhanced search functionality"""
    
    def get_search_fields(self):
        """Get fields to search in"""
        return getattr(self, 'search_fields', [])
    
    def filter_search(self, queryset, search_query):
        """Filter queryset by search query"""
        if not search_query:
            return queryset
        
        search_fields = self.get_search_fields()
        if not search_fields:
            return queryset
        
        query = Q()
        for field in search_fields:
            if '__' in field:  # Related field
                query |= Q(**{f"{field}__icontains": search_query})
            else:
                query |= Q(**{f"{field}__icontains": search_query})
        
        return queryset.filter(query)